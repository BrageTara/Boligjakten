from playwright.sync_api import sync_playwright
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime
import time
import re
import os
import json

EXCEL_FILE = "finn_tracker.xlsx"

# Trondheim postnummer → nabolag
POSTNUMMER = {
    # Sentrum / indre by
    "7010": "Midtbyen",       "7011": "Midtbyen",       "7012": "Singsaker",
    "7013": "Øya",            "7014": "Midtbyen",       "7015": "Tempe",
    "7016": "Midtbyen",       "7017": "Midtbyen",       "7018": "Ila",
    "7019": "Midtbyen",       "7020": "Trondheim",      "7021": "Byåsen",
    "7022": "Byåsen",         "7023": "Nardo",          "7024": "Stavne",
    "7025": "Brundalen",      "7026": "Trondheim",      "7027": "Lademoen",
    "7028": "Bromstad",       "7029": "Trondheim",      "7030": "Elgeseter",
    "7031": "Elgeseter",      "7032": "Trondheim",      "7033": "Risvollan",
    "7034": "Moholt",         "7035": "Trondheim",      "7036": "Trondheim",
    "7037": "Rosenborg",      "7038": "Trondheim",      "7039": "Trondheim",
    "7040": "Trondheim",      "7041": "Nedre Lade",     "7042": "Øvre Lade",
    "7043": "Møllenberg",     "7044": "Nedre Elvehavn", "7045": "Trolla",
    "7046": "Trondheim",      "7047": "Trondheim",      "7048": "Strindheim",
    "7049": "Vikåsen",        "7050": "Trondheim",      "7051": "Charlottenlund",
    "7052": "Trondheim",
    # Øst / Ranheim
    "7053": "Ranheim",        "7054": "Ranheim",        "7055": "Ranheim",
    "7056": "Ranheim",        "7057": "Jonsvatnet",     "7058": "Jakobsli",
    "7059": "Jakobsli",
    # Klæbu (del av Trondheim kommune fra 2020)
    "7060": "Klæbu",          "7066": "Trondheim",      "7067": "Trondheim",
    "7068": "Trondheim",      "7069": "Trondheim",
    # Sør / Heimdal / Tiller
    "7070": "Bosberg",        "7071": "Trondheim",      "7072": "Heimdal",
    "7074": "Spongdal",       "7075": "Tiller",         "7078": "Saupstad",
    "7079": "Flatåsen",       "7080": "Heimdal",        "7081": "Sjetnemarka",
    "7082": "Kattem",         "7083": "Leinstrand",     "7088": "Heimdal",
    "7089": "Heimdal",        "7091": "Tiller",         "7092": "Tiller",
    "7093": "Tiller",         "7097": "Saupstad",       "7098": "Saupstad",
    "7099": "Flatåsen",
    # Klæbu tettsted
    "7540": "Klæbu",          "7541": "Klæbu",          "7548": "Tanem",
    "7549": "Tanem",
}

SEARCH_URL = (
    "https://www.finn.no/realestate/homes/search.html"
    "?location=1.20016.20318&property_type=3"
    "&sort=PUBLISHED_DESC"
)

SEARCH_DELAY = 3   # sekunder mellom søkesider
AD_DELAY     = 2   # sekunder mellom annonsesider

COLS = {
    "finnkode":         ("A", "Finnkode"),
    "adresse":          ("B", "Adresse"),
    "prisantydning":    ("C", "Prisantydning"),
    "fellesgjeld":      ("D", "Fellesgjeld"),
    "totalpris":        ("E", "Totalpris"),
    "kvm_pris":         ("F", "kr/m²"),
    "felleskost":       ("G", "Felleskost/mnd"),
    "fellesformue":     ("H", "Fellesformue"),
    "type":             ("I", "Boligtype"),
    "bra":              ("J", "BRA (m²)"),
    "rom":              ("K", "Rom"),
    "etasje":           ("L", "Etasje"),
    "forste_sett":      ("M", "Første gang sett"),
    "siste_sett":       ("N", "Sist oppdatert"),
    "dager_ute":        ("O", "Dager på Finn"),
    "antall_visninger": ("P", "Antall visninger"),
    "pris_ved_start":   ("Q", "Startpris"),
    "prisendring":      ("R", "Prisendring"),
    "status":           ("S", "Status"),
    "url":              ("T", "URL"),
    "megler":           ("U", "Megler"),
    "meglerkontor":     ("V", "Meglerkontor"),
    "neste_visning":    ("W", "Neste visning"),
    "flagg":            ("X", "Flagg"),
    "omrade":           ("Y", "Område"),
    "postnummer":       ("Z", "Postnummer"),
}

BLUE  = Font(color="0000FF")
BLACK = Font(color="000000")
GREEN = Font(color="008000")
BOLD  = Font(bold=True, color="000000")
HEADER_FILL = PatternFill("solid", start_color="D9D9D9")
SOLD_FILL   = PatternFill("solid", start_color="FFD7D7")
NEW_FILL    = PatternFill("solid", start_color="D7FFD7")
CENTER = Alignment(horizontal="center")
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


# ─── Playwright scraping ───────────────────────────────────────────────────────

def accept_cookies(page):
    """Godta cookie-dialogen hvis den dukker opp."""
    try:
        btn = page.query_selector('button:has-text("Godta alle")')
        if btn:
            btn.click()
            print("  Cookie-dialog godtatt.")
            time.sleep(1)
    except Exception:
        pass


def fetch_all_listings(page):
    """Henter alle finnkoder fra søkeresultatsidene med Playwright."""
    all_listings = []
    seen = set()
    page_num = 1

    while True:
        url = SEARCH_URL if page_num == 1 else f"{SEARCH_URL}&page={page_num}"
        print(f"  Henter søkeside {page_num}...")

        # Retry ved nettverksfeil
        for attempt in range(3):
            try:
                page.goto(url, wait_until="domcontentloaded")
                time.sleep(1)
                accept_cookies(page)
                page.wait_for_selector("article", timeout=15000)
                break
            except Exception as e:
                err = str(e)
                if "ERR_NAME_NOT_RESOLVED" in err or "ERR_INTERNET" in err or "ERR_NETWORK" in err:
                    if attempt < 2:
                        print(f"    Nettverksfeil, prøver igjen om 10s...")
                        time.sleep(10)
                    else:
                        raise
                elif "Timeout" in err and "article" in err:
                    # Ingen annonser på denne siden — ferdig med søk
                    print(f"  Side {page_num} har ingen annonser — ferdig med søk.")
                    return all_listings
                else:
                    raise

        time.sleep(1)  # la eventuelle lazy-load elementer laste

        # Hent alle annonselenker fra rendret DOM
        links = page.eval_on_selector_all(
            'article a[href*="/realestate/"]',
            """elements => elements.map(el => el.href).filter(h => h.includes('/ad.html') || h.includes('/newbuildings/'))"""
        )

        batch = []
        for href in links:
            m = re.search(r"finnkode=(\d+)", href)
            if not m:
                m = re.search(r"/(\d+)(?:\?|$)", href)
            if m:
                finnkode = m.group(1)
                if finnkode not in seen:
                    seen.add(finnkode)
                    batch.append({
                        "finnkode": finnkode,
                        "url": href,
                    })

        if not batch:
            break

        all_listings.extend(batch)
        page_num += 1
        time.sleep(SEARCH_DELAY)

    print(f"  Totalt hentet {page_num - 1} søkesider.")
    return all_listings


def scrape_ad(page, url):
    """Scraper en enkelt annonse med Playwright."""
    try:
        page.goto(url, wait_until="domcontentloaded")
        page.wait_for_selector("h1", timeout=15000)
        time.sleep(1)
    except Exception:
        return {}

    data = {}
    html = page.content()

    def find_val(label_text):
        """Finn verdi i dt/dd eller th/td-par."""
        try:
            # Prøv dt/dd
            dts = page.query_selector_all("dt")
            for dt in dts:
                if label_text.lower() in dt.inner_text().lower():
                    dd = dt.evaluate("el => el.nextElementSibling?.textContent?.trim()")
                    if dd:
                        return dd
            # Prøv th/td
            ths = page.query_selector_all("th")
            for th in ths:
                if label_text.lower() in th.inner_text().lower():
                    td = th.evaluate("el => el.nextElementSibling?.textContent?.trim()")
                    if td:
                        return td
        except Exception:
            pass
        return None

    def parse_kr(text):
        if not text:
            return None
        digits = re.sub(r"[^\d]", "", text)
        return int(digits) if digits else None

    # Adresse fra h1
    try:
        h1 = page.query_selector("h1")
        data["adresse"] = h1.inner_text().strip() if h1 else None
    except Exception:
        data["adresse"] = None

    # Prisantydning
    try:
        price_el = page.query_selector('[data-testid="pricing-incicative-price"]')
        if price_el:
            spans = price_el.query_selector_all("span")
            data["prisantydning"] = parse_kr(spans[1].inner_text()) if len(spans) >= 2 else None
        else:
            data["prisantydning"] = None
    except Exception:
        data["prisantydning"] = None

    data["fellesgjeld"]   = parse_kr(find_val("Fellesgjeld"))
    data["felleskost"]    = parse_kr(find_val("Felleskost"))
    data["fellesformue"]  = parse_kr(find_val("Fellesformue"))
    data["type"]          = find_val("Boligtype")
    data["bra"]           = find_val("Internt bruksareal") or find_val("Bruksareal") or find_val("Primærrom")
    data["rom"]           = find_val("Soverom")
    data["etasje"]        = find_val("Etasje")

    # Totalpris
    totalpris_raw = parse_kr(find_val("Totalpris"))
    if totalpris_raw:
        data["totalpris"] = totalpris_raw
    elif data["prisantydning"] and data["fellesgjeld"]:
        data["totalpris"] = data["prisantydning"] + data["fellesgjeld"]
    elif data["prisantydning"]:
        data["totalpris"] = data["prisantydning"]
    else:
        data["totalpris"] = None

    # Megler
    data["megler"] = None
    data["meglerkontor"] = None
    m_org = re.search(r'"orgName"\s*:\s*"([^"]+)"', html)
    m_name = re.search(r'"name"\s*:\s*"([^"]+)"', html)
    if m_org:
        data["meglerkontor"] = m_org.group(1)
    if m_name:
        data["megler"] = m_name.group(1)

    # Postnummer
    data["postnummer"] = None
    pnr = None

    pnr_raw = find_val("Postnummer") or find_val("postnummer")
    if pnr_raw:
        m = re.search(r"\b(\d{4})\b", pnr_raw)
        if m:
            pnr = m.group(1)

    if not pnr:
        for field in ("postCode", "zipCode", "zip_code", "postal_code",
                      "postalCode", "postcode", "zip"):
            m = re.search(rf'"{field}"\s*:\s*"?(\d{{4}})"?', html)
            if m:
                pnr = m.group(1)
                break

    if not pnr:
        m = re.search(r"\b(\d{4})\s+[A-ZÆØÅ]{2,}", html)
        if m:
            pnr = m.group(1)

    if not pnr:
        m = re.search(r"\b(\d{4})\s+[A-ZÆØÅ][a-zA-ZæøåÆØÅ]+", html)
        if m:
            pnr = m.group(1)

    if pnr and not (1000 <= int(pnr) <= 9999):
        pnr = None

    if pnr:
        data["postnummer"] = pnr

    data["omrade"] = POSTNUMMER.get(pnr, f"Ukjent ({pnr})") if pnr else None

    # Neste visning
    data["neste_visning"] = None
    try:
        viewing_el = page.query_selector('[data-testid^="viewings-"]')
        if viewing_el:
            date_el = viewing_el.query_selector(".capitalize-first")
            time_el = viewing_el.query_selector("[class*='font-bold']")
            if date_el and time_el:
                dato = date_el.inner_text().strip()
                tid = re.sub(r"\s+", " ", time_el.inner_text()).strip()
                data["neste_visning"] = f"{dato} {tid}"
    except Exception:
        pass

    return data


def check_sold_status(page, url):
    """Besøker annonsesiden og returnerer årsaken til at den er borte fra søk."""
    try:
        resp = page.goto(url, wait_until="domcontentloaded")
    except Exception:
        return "Ukjent"

    time.sleep(1)
    html = page.content()
    text_lower = html.lower()

    if resp and resp.status == 404:
        return "Trukket"

    if "siden finnes ikke" in text_lower or "page not found" in text_lower:
        return "Trukket"

    if "annonsen er ikke lenger tilgjengelig" in text_lower or "ikke lenger aktiv" in text_lower:
        return "Trukket"

    # Prøv React Router stream-dekoding
    stream_result = _parse_stream_sold_state(html)
    if stream_result is True:
        return "Solgt"
    if stream_result is False:
        return "Trukket"

    if ("boligen er solgt" in text_lower or ">solgt<" in text_lower
            or '"sold":true' in html or '"isSold":true' in html):
        return "Solgt"

    return "Ukjent"


def _parse_stream_sold_state(html_text):
    """
    Parser React Router stream-data fra Finn.no for å sjekke realEstateSoldState.
    Returnerer True (solgt), False (aktiv) eller None (ukjent).
    """
    marker = "streamController.enqueue("
    idx = html_text.find(marker)
    if idx < 0:
        return None
    content_start = idx + len(marker)
    raw_str = html_text[content_start:]
    end_idx = raw_str.find(");</script>")
    if end_idx < 0:
        return None
    js_literal = raw_str[:end_idx]
    try:
        inner = json.loads(js_literal)
        arr = json.loads(inner)
    except Exception:
        return None

    try:
        sold_key_idx = arr.index("realEstateSoldState")
    except ValueError:
        return None

    sold_key_str = f"_{sold_key_idx}"

    for item in arr:
        if isinstance(item, dict) and sold_key_str in item:
            val = item[sold_key_str]
            if isinstance(val, int):
                return val >= 0
    return None


# ─── Excel-logikk (uendret) ───────────────────────────────────────────────────

SOLD_EXTRA_COLS = [
    ("Z",  "Solgt dato"),
    ("AA", "Årsak"),
]


def init_sold_sheet(wb):
    sold_ws = wb.create_sheet("Solgte")
    sold_ws.freeze_panes = "A2"
    sold_ws.auto_filter.ref = f"A1:AA1"

    for key, (col, label) in COLS.items():
        cell = sold_ws[f"{col}1"]
        cell.value = label
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    for col, label in SOLD_EXTRA_COLS:
        cell = sold_ws[f"{col}1"]
        cell.value = label
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    widths = {
        "A": 12, "B": 35, "C": 16, "D": 14, "E": 16,
        "F": 14, "G": 16, "H": 14, "I": 14, "J": 10,
        "K": 8,  "L": 10, "M": 16, "N": 16, "O": 12,
        "P": 16, "Q": 16, "R": 14, "S": 12, "T": 50,
        "U": 22, "V": 30, "W": 24, "X": 30, "Y": 18, "Z": 14, "AA": 14,
    }
    for col, w in widths.items():
        sold_ws.column_dimensions[col].width = w

    return sold_ws


def migrate_sold_columns(sold_ws):
    existing_headers = {sold_ws.cell(1, c).value: c for c in range(1, sold_ws.max_column + 1)}
    changed = False
    for col_letter, label in SOLD_EXTRA_COLS:
        if label not in existing_headers:
            col_num = (
                ord(col_letter) - ord("A") + 1
                if len(col_letter) == 1
                else 26 + ord(col_letter[1]) - ord("A") + 1
            )
            cell = sold_ws.cell(1, col_num)
            cell.value = label
            cell.font = BOLD
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = BORDER
            sold_ws.column_dimensions[col_letter].width = 14
            print(f"  Ny kolonne i Solgte: {col_letter} — {label}")
            changed = True
    return changed


def init_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Annonser"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    for key, (col, label) in COLS.items():
        cell = ws[f"{col}1"]
        cell.value = label
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    widths = {
        "A": 12, "B": 35, "C": 16, "D": 14, "E": 16,
        "F": 14, "G": 16, "H": 14, "I": 14, "J": 10,
        "K": 8,  "L": 10, "M": 16, "N": 16, "O": 12,
        "P": 16, "Q": 16, "R": 14, "S": 12, "T": 50,
        "U": 22, "V": 30, "W": 24, "X": 30, "Y": 18, "Z": 12,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    init_sold_sheet(wb)

    log_ws = wb.create_sheet("Prishistorikk")
    for col, label in [("A", "Finnkode"), ("B", "Dato"), ("C", "Prisantydning"), ("D", "Totalpris")]:
        cell = log_ws[f"{col}1"]
        cell.value = label
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.border = BORDER
    for col, w in [("A", 12), ("B", 14), ("C", 16), ("D", 16)]:
        log_ws.column_dimensions[col].width = w

    wb.save(EXCEL_FILE)
    print(f"Opprettet {EXCEL_FILE}")


def migrate_columns(ws):
    existing_headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    changed = False
    for key, (col_letter, label) in COLS.items():
        if label not in existing_headers:
            col_num = ord(col_letter) - ord("A") + 1
            cell = ws.cell(1, col_num)
            cell.value = label
            cell.font = BOLD
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = BORDER
            ws.column_dimensions[col_letter].width = 16
            print(f"  Ny kolonne lagt til: {col_letter} — {label}")
            changed = True
    return changed


def get_header_map(ws):
    return {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}


def load_existing():
    if not os.path.exists(EXCEL_FILE):
        init_workbook()
    wb = load_workbook(EXCEL_FILE)
    if "Solgte" not in wb.sheetnames:
        init_sold_sheet(wb)
        wb.save(EXCEL_FILE)
    ws = wb["Annonser"]
    sold_ws = wb["Solgte"]
    needs_save = migrate_columns(ws)
    needs_save |= migrate_sold_columns(sold_ws)
    if needs_save:
        wb.save(EXCEL_FILE)

    header_map = get_header_map(ws)
    idx = {label: col - 1 for label, col in header_map.items()}

    existing = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        fk_idx = idx.get("Finnkode", 0)
        if row[fk_idx]:
            finnkode = str(row[fk_idx])

            def get(label, row=row):
                i = idx.get(label)
                return row[i] if i is not None and i < len(row) else None

            existing[finnkode] = {
                "row_data": row,
                "forste_sett": get("Første gang sett"),
                "pris_ved_start": get("Startpris"),
                "neste_visning": get("Neste visning"),
                "antall_visninger": get("Antall visninger") or 0,
            }
    return wb, existing


def write_row(ws, row_num, finnkode, ad, today, existing_info=None):
    forste_sett = today
    pris_ved_start = ad.get("prisantydning")
    is_new = True

    antall_visninger = 0
    if existing_info:
        forste_sett = existing_info.get("forste_sett") or today
        pris_ved_start = existing_info.get("pris_ved_start") or ad.get("prisantydning")
        is_new = False
        antall_visninger = existing_info.get("antall_visninger") or 0
        forrige_visning = existing_info.get("neste_visning")
        ny_visning = ad.get("neste_visning")
        if ny_visning and ny_visning != forrige_visning:
            antall_visninger += 1

    if isinstance(forste_sett, datetime):
        forste_sett = forste_sett.date()
    dager = (today - forste_sett).days if isinstance(forste_sett, date) else 0

    prisendring = None
    if pris_ved_start and ad.get("prisantydning"):
        prisendring = ad["prisantydning"] - pris_ved_start

    kvm_pris = None
    bra_str = ad.get("bra")
    if bra_str and ad.get("prisantydning"):
        bra_match = re.search(r"(\d+)", str(bra_str))
        if bra_match:
            bra_num = int(bra_match.group(1))
            if bra_num > 0:
                kvm_pris = round(ad["prisantydning"] / bra_num)

    flagg = []
    if prisendring and prisendring < 0:
        flagg.append("Prisnedsatt")
    if dager >= 14:
        flagg.append("14+ dager")
    if antall_visninger >= 2:
        flagg.append("2+ visninger")
    if ad.get("totalpris") and ad.get("fellesgjeld") and ad["fellesgjeld"] / ad["totalpris"] >= 0.30:
        flagg.append("Høy fellesgjeld")

    label_values = {
        "Finnkode":          finnkode,
        "Adresse":           ad.get("adresse"),
        "Prisantydning":     ad.get("prisantydning"),
        "Fellesgjeld":       ad.get("fellesgjeld"),
        "Totalpris":         ad.get("totalpris"),
        "kr/m²":             kvm_pris,
        "Felleskost/mnd":    ad.get("felleskost"),
        "Fellesformue":      ad.get("fellesformue"),
        "Boligtype":         ad.get("type"),
        "BRA (m²)":          ad.get("bra"),
        "Rom":               ad.get("rom"),
        "Etasje":            ad.get("etasje"),
        "Første gang sett":  forste_sett,
        "Sist oppdatert":    today,
        "Dager på Finn":     dager,
        "Antall visninger":  antall_visninger if antall_visninger else None,
        "Startpris":         pris_ved_start,
        "Prisendring":       prisendring,
        "Status":            "Aktiv",
        "URL":               f"https://www.finn.no/realestate/homes/ad.html?finnkode={finnkode}",
        "Megler":            ad.get("megler"),
        "Meglerkontor":      ad.get("meglerkontor"),
        "Neste visning":     ad.get("neste_visning"),
        "Flagg":             " | ".join(flagg) if flagg else None,
        "Område":            ad.get("omrade"),
        "Postnummer":        ad.get("postnummer"),
    }

    header_map = get_header_map(ws)
    for label, value in label_values.items():
        col_num = header_map.get(label)
        if col_num is None:
            continue
        cell = ws.cell(row_num, col_num, value)
        cell.border = BORDER
        if label == "URL":
            cell.hyperlink = str(value) if value else None
            cell.font = BLUE
        elif label == "Flagg" and value:
            cell.font = Font(bold=True, color="CC0000")
        elif label == "Prisendring" and value and value < 0:
            cell.font = GREEN
        else:
            cell.font = BLACK

        if label in ("Prisantydning", "Fellesgjeld", "Totalpris", "kr/m²",
                     "Felleskost/mnd", "Fellesformue", "Startpris"):
            cell.number_format = '#,##0'
        elif label == "Prisendring":
            cell.number_format = '+#,##0" kr";"-"#,##0" kr";"-"'
        elif label in ("Første gang sett", "Sist oppdatert"):
            cell.number_format = 'YYYY-MM-DD'

    if is_new:
        for col_num in range(1, len(label_values) + 1):
            ws.cell(row_num, col_num).fill = NEW_FILL


def log_price_history(wb, finnkode, today, prisantydning, totalpris):
    if "Prishistorikk" not in wb.sheetnames:
        return
    log_ws = wb["Prishistorikk"]
    next_row = log_ws.max_row + 1
    log_ws.cell(next_row, 1, finnkode)
    log_ws.cell(next_row, 2, today).number_format = 'YYYY-MM-DD'
    if prisantydning:
        log_ws.cell(next_row, 3, prisantydning).number_format = '#,##0'
    if totalpris:
        log_ws.cell(next_row, 4, totalpris).number_format = '#,##0'


def mark_sold(wb, finnkode, row_data, today, arsak="Ukjent"):
    if "Solgte" not in wb.sheetnames:
        init_sold_sheet(wb)
    sold_ws = wb["Solgte"]
    next_row = sold_ws.max_row + 1
    for col_num, value in enumerate(row_data, start=1):
        cell = sold_ws.cell(next_row, col_num, value)
        cell.fill = SOLD_FILL
        cell.border = BORDER
        cell.font = BLACK
    sold_ws.cell(next_row, 26, today).number_format = 'YYYY-MM-DD'
    sold_ws.cell(next_row, 27, arsak)
    for col_num in range(1, 28):
        sold_ws.cell(next_row, col_num).fill = SOLD_FILL


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    today = date.today()
    print(f"\n=== Finn-tracker kjøres {today} ===\n")

    wb, existing = load_existing()
    ws = wb["Annonser"]

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            )
        )
        pg = context.new_page()

        print("Henter søkeresultater fra Finn.no...")
        try:
            listings = fetch_all_listings(pg)
        except Exception as e:
            print(f"FEIL ved henting av søkeside: {e}")
            browser.close()
            return

        print(f"  Fant {len(listings)} annonser i søket.\n")

        found_codes = {l["finnkode"] for l in listings}
        new_count = 0
        updated_count = 0

        header_map = get_header_map(ws)
        fk_col = header_map.get("Finnkode", 1)
        row_map = {}
        for row_num in range(2, ws.max_row + 1):
            val = ws.cell(row_num, fk_col).value
            if val:
                row_map[str(val)] = row_num

        try:
            for i, listing in enumerate(listings, 1):
                finnkode = listing["finnkode"]
                url = listing["url"]
                print(f"  [{i}/{len(listings)}] Henter {finnkode}...", end=" ", flush=True)
                ad = scrape_ad(pg, url)
                if not ad:
                    print("ingen data, hopper over.")
                    continue

                status = "ny" if finnkode not in row_map else "oppdaterer"
                print(f"{status}  |  {ad.get('adresse', '')}  |  postnr={ad.get('postnummer')}  område={ad.get('omrade')}")

                if finnkode in row_map:
                    row_num = row_map[finnkode]
                    write_row(ws, row_num, finnkode, ad, today, existing_info=existing.get(finnkode))
                    updated_count += 1
                else:
                    row_num = ws.max_row + 1
                    write_row(ws, row_num, finnkode, ad, today, existing_info=None)
                    row_map[finnkode] = row_num
                    new_count += 1

                log_price_history(wb, finnkode, today, ad.get("prisantydning"), ad.get("totalpris"))
                time.sleep(AD_DELAY)

            # Sjekk om noen eksisterende annonser er borte fra søket
            sold_count = 0
            for finnkode, info in existing.items():
                if finnkode not in found_codes:
                    row_num = row_map.get(finnkode)
                    if row_num:
                        status_col = header_map.get("Status")
                        if status_col and ws.cell(row_num, status_col).value in ("Aktiv", "Ukjent"):
                            ad_url = f"https://www.finn.no/realestate/homes/ad.html?finnkode={finnkode}"
                            arsak = check_sold_status(pg, ad_url)
                            if status_col:
                                ws.cell(row_num, status_col).value = arsak
                                ws.cell(row_num, status_col).font = Font(bold=True, color="CC0000")
                            print(f"  {finnkode} ikke lenger i søk — status: {arsak}")
                            mark_sold(wb, finnkode, info["row_data"], today, arsak)
                            sold_count += 1

        finally:
            browser.close()
            # Lagre — fallback til backup hvis filen er låst
            save_path = EXCEL_FILE
            try:
                wb.save(save_path)
            except (IOError, PermissionError):
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                save_path = EXCEL_FILE.replace(".xlsx", f"_backup_{ts}.xlsx")
                wb.save(save_path)
                print(f"\nADVARSEL: {EXCEL_FILE} var låst (åpen i Excel?). Lagret til {save_path} i stedet.")

            print(f"\n=== Ferdig ===")
            print(f"  Nye annonser:     {new_count}")
            print(f"  Oppdaterte:       {updated_count}")
            print(f"  Solgt/trukket:    {sold_count}")
            print(f"  Lagret til:       {save_path}\n")


if __name__ == "__main__":
    main()
